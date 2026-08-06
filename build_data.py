"""
Build fauna.json + flora.json otomatis
Jalankan: python build_data.py
"""
import json, os, glob
from openpyxl import load_workbook

# =================== FAUNA ===================
print("=== MEMBANGUN FAUNA.JSON ===")
wb = load_workbook("Data burung saukabu.xlsx")
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cols = (list(row) + [None]*8)[:8]
    no, nama_ilmiah, nama_lokal, famili, iucn, cites, p106, _ = cols
    if not nama_ilmiah: continue
    rows.append({
        "no": no,
        "nama_ilmiah": str(nama_ilmiah).strip(),
        "nama_lokal": str(nama_lokal).strip() if nama_lokal else "",
        "famili": str(famili).strip() if famili else "",
        "iucn": str(iucn).strip() if iucn else "LC",
        "cites": str(cites).strip() if cites else "-",
        "p106": str(p106).strip() if p106 else "-",
    })
print(f"  {len(rows)} spesies di Excel")

# Cari foto: prefer .jpeg (HQ), fallback .jpg
fauna_dir = "images/fauna"
def cari_foto(nama_ilmiah):
    key = nama_ilmiah.replace(" ", "_")
    for ext in ["jpeg", "jpg", "png", "webp"]:
        p = f"{fauna_dir}/{key}.{ext}"
        if os.path.exists(p): return p
    return ""

# Deskripsi per spesies
DESC = {
    "Eclectus roratus": "Burung Bayan atau Nuri Kepala Biru adalah salah satu spesies paling mencolok di Papua. Jantan berwarna hijau cerah dengan paruh merah-oranye; betina berwarna merah tua dengan perut biru dan paruh hitam — perbedaan ekstrem hingga sempat dikira dua spesies berbeda. Sering terlihat berpasangan di tajuk pohon tinggi sekitar Saukabu.",
    "Geoffroyus geoffroyi": "Nuri Pipi-Merah adalah nuri berukuran sedang yang sangat aktif. Jantan berciri kepala merah muda keunguan yang khas, betina berkepala cokelat. Keduanya bertubuh hijau dominan. Spesies ini berasosiasi kuat dengan hutan primer yang sehat dan sering terlihat terbang cepat di antara kanopi pohon buah di Saukabu.",
    "Ptilinopus rivoli": "Walik Dada Putih adalah merpati buah elegan di Raja Ampat. Jantan berwarna hijau zamrud berkilau dengan dada putih bersih yang kontras mencolok. Spesies ini menghuni hutan primer dan tepi hutan, memakan aneka buah dari pohon tinggi. Di Saukabu, suara panggilannya yang berulang menjadi penanda kehadiran hutan sehat.",
    "Rhipidura leucophrys": "Kipasan Belang adalah burung kecil yang aktif dan tidak takut manusia. Dengan bulu hitam-putih kontras dan ekor yang selalu dikibas-kibaskan, ia menjadi teman setia saat berjalan di sekitar kampung dan tepi pantai. Kehadirannya menjadi indikator ekosistem yang masih terjaga.",
    "Tanysiptera galatea": "Cekakak Sega adalah kingfisher bermahkota biru cerah dengan ekor panjang berwarna putih yang sangat khas. Spesies ikonik Raja Ampat ini menghuni hutan dataran rendah dan menjadi daya tarik utama bagi pengamat burung dari seluruh dunia.",
    "Todiramphus sanctus": "Cekakak Suci adalah kingfisher berukuran sedang dengan punggung biru-hijau mengkilap dan dada putih bersih. Spesies migran parsial ini sering terlihat di tepi pantai, mangrove, dan kebun kelapa di sekitar Saukabu.",
    "Todiramphus saurophagus": "Cekakak Raksasa adalah kingfisher terbesar di kawasan Papua. Dengan tubuh besar, paruh kuat, dan bulu biru-putih yang mencolok, spesies ini mudah dikenali. Ia memangsa kadal, ular kecil, dan crustasea di tepi pantai dan hutan mangrove.",
    "Cinnyris jugularis": "Burung Madu Sriganti adalah burung madu yang paling umum di Asia Tenggara dan Papua. Jantan memiliki bulu atas hijau mengkilap dan tenggorokan merah-ungu yang indah. Sangat aktif mengunjungi pohon-pohon berbunga di sekitar kampung.",
    "Megapodius freycinet": "Maleo Gosong adalah megapoda unik yang tidak mengerami telurnya sendiri, melainkan mengubur telur di dalam gundukan pasir hangat. Panas bumi dan matahari menggantikan peran induk. Spesies ini adalah insinyur alam yang luar biasa.",
    "Egretta sacra": "Kuntul Karang hidup di ekosistem pesisir berbatu. Terdapat dua morfa: putih seluruhnya atau abu-abu gelap seluruhnya. Ia adalah predator aktif ikan dan crustasea di zona pasang surut, sering terlihat berdiri diam di atas batu karang.",
    "Ardea alba": "Kuntul Besar adalah kuntul putih berukuran besar yang umum dijumpai di kawasan pesisir dan rawa. Ia memangsa ikan, katak, dan serangga di perairan dangkal. Di Saukabu, terlihat mencari makan saat air surut di pantai.",
    "Fregata ariel": "Cikalang Kecil adalah burung laut dengan kemampuan terbang luar biasa. Seperti cikalang lainnya, ia tidak bisa mendarat di air karena bulunya tidak tahan air. Sering merampas makanan dari burung lain (kleptoparasitisme) di atas perairan Saukabu.",
    "Fregata minor": "Cikalang Besar memiliki rentang sayap lebih dari dua meter dan merupakan penguasa udara di atas perairan tropis. Kantung gular merah jantan menjadi atraksi musim kawin yang menakjubkan. Terlihat melayang tinggi di atas laut sekitar Piaynemo.",
    "Haliastur indus": "Elang Bondol adalah elang pesisir dengan tubuh putih bersih dan sayap cokelat kemerahan. Sangat mudah dijumpai di seluruh nusantara. Di Saukabu, sering terlihat melayang di atas pantai mencari ikan dan crustasea.",
    "Haliaeetus leucogaster": "Elang Laut Perut Putih adalah salah satu elang terbesar di kawasan Asia-Pasifik. Dengan bentang sayap hingga 2,2 meter dan warna putih-abu-abu yang khas, kemunculannya di langit Saukabu selalu menjadi pemandangan yang mengesankan.",
    "Accipiter hiogaster": "Elang Alap Papua adalah accipiter berukuran sedang yang khas Papua. Spesies ini merupakan predator burung-burung kecil dan reptil di hutan dataran rendah. Penerbangannya yang cepat dan lincah di antara pepohonan menjadikannya pemburu yang sangat efektif.",
    "Lophospiza trivirgata": "Elang Alap Sulawesi adalah raptor berukuran kecil-sedang yang juga tercatat di Papua bagian barat. Ia menghuni hutan primer dan sekunder yang masih baik, memangsa burung kecil dan serangga besar.",
    "Esacus magnirostris": "Wili-wili Besar adalah burung pesisir berbadan besar dengan paruh yang tebal dan kuat. Spesies ini khas di pantai berbatu dan terumbu karang, aktif di malam hari mencari kepiting dan moluska. Terlihat di pantai berbatu Pulau Fam.",
    "Tadorna radjah": "Itik Radjah atau Angsa Radjah adalah bebek besar yang khas dengan bulu putih, kepala merah muda, dan paruh merah. Spesies endemik Papua-Australia ini menghuni hutan mangrove dan tepi sungai dekat pantai.",
    "Megapodius freycinet": "Gosong Kelam adalah megapoda yang membangun gundukan sarang dari material organik yang membusuk untuk menghasilkan panas penetasan. Spesies ini menjadi indikator penting ekosistem hutan tropis yang masih sehat.",
    "Macropygia doreya": "Uncal Muda adalah merpati bermata merah dan berwarna cokelat kemerahan. Ia menyukai tepi hutan, semak belukar, dan kebun di kawasan Papua. Sering terdengar berkukuk lembut di pagi hari dari balik pepohonan di sekitar kampung.",
    "Ducula aenea": "Pergam Hijau adalah merpati besar pemakan buah yang penting bagi regenerasi hutan. Dengan bulu hijau mengkilap dan perut merah muda, spesies ini memiliki peran krusial sebagai pemencar biji pohon buah-buahan hutan.",
    "Ducula myristicivora": "Pergam Pala adalah spesies pergam yang sangat menyukai buah pala. Bulu putih bersih dengan bagian atas abu-abu menjadikannya mudah dikenali. Sangat penting bagi ekologi hutan karena menyebarkan biji pala ke seluruh kawasan.",
    "Eos squamata": "Nuri Kalung-Ungu atau Kakatua Raja Irian adalah lori berwarna merah dengan detail ungu dan biru yang memukau. Spesies endemik Maluku dan Papua Barat ini sering terlihat berkelompok besar di pohon berbunga, memakan nektar dan serbuk sari.",
    "Cacatua galerita": "Kakatua Putih Besar atau Kakatua Jambul Kuning adalah kakatua yang paling ikonik. Dengan bulu putih bersih dan jambul kuning yang spektakuler, spesies ini sangat cerdas. Sering terlihat berkelompok berisik di tajuk pohon tinggi.",
    "Caprimulgus macrurus": "Cabak Maling adalah burung malam dari famili Caprimulgidae yang aktif saat senja dan malam hari. Dengan pola bulu yang menyerupai kulit kayu, ia sempurna berkamuflase di lantai hutan saat beristirahat di siang hari.",
    "Eurystomus orientalis": "Tiong Batu atau Dollarbird adalah burung berwarna biru-hijau mengkilap dengan paruh merah dan bercak putih berbentuk koin di sayap saat terbang. Ia berburu serangga besar di udara dari tenggeran yang tinggi.",
    "Merops ornatus": "Kirik-kirik Australia adalah burung yang cantik dengan kombinasi warna hijau, biru, kuning, dan merah. Ia adalah pemangsa lebah dan serangga lain yang ditangkap di udara. Spesies migran ini singgah di Papua saat musim non-breeding.",
    "Coracina papuensis": "Kepudang Sungu Papua adalah burung dari famili Campephagidae dengan bulu abu-abu pucat dan sayap hitam. Ia menghuni tajuk hutan dataran rendah dan sering bergabung dengan kelompok burung campuran.",
    "Hirundo tahitica": "Layang-layang Pasifik adalah layang-layang dengan punggung biru-hijau metalik dan perut putih. Sangat umum di kawasan pesisir dan pemukiman. Di Saukabu, terlihat terbang berputar rendah di atas pantai saat berburu serangga.",
    "Pachycephala simplex": "Kancilan Polos adalah burung penyanyi dari famili Pachycephalidae dengan tubuh ramping dan warna yang polos. Meski penampilannya sederhana, suara nyanyiannya yang merdu menjadikannya mudah didengar di hutan sekitar Saukabu.",
    "Myiagra alecto": "Kehicap Ratu adalah flycatcher berukuran kecil-sedang. Jantan berwarna hitam mengkilap metalik, betina lebih pucat. Spesies ini aktif memangsa serangga di lapisan tengah hutan dan sering terlihat di kebun-kebun warga.",
    "Artamus leucorynchus": "Kekep Babi adalah burung kelabu-putih berukuran sedang yang sering terlihat bertengger berkelompok di kabel listrik atau ranting pohon mati. Ia memangsa serangga yang ditangkap di udara dengan manuver terbang yang lincah.",
    "Gavicalis versicolor": "Isap Madu Berbintik adalah burung madu dari Australia yang juga tercatat di Papua. Dengan paruh melengkung dan lidah khusus untuk menghisap nektar, ia berperan penting sebagai penyerbuk bunga-bunga tropis.",
    "Aplonis metallica": "Perling Sutera adalah jalak dengan bulu hitam mengkilap berminyak metalik yang mencolok. Hidup berkoloni besar, membangun sarang menggantung ratusan buah dalam satu pohon. Di Saukabu, koloninya menjadi pemandangan yang sangat ramai di sore hari.",
    "Corvus orru": "Gagak Torresian adalah gagak hitam legam yang sangat cerdas dan adaptif. Ia adalah omnivora oportunistik yang memakan hampir segalanya. Di Saukabu, sering terlihat berkelompok dan berinteraksi dengan aktivitas warga.",
    "Dicrurus bracteatus": "Srigunting Perahu adalah burung hitam mengkilap dengan ekor bercabang seperti garpu yang khas. Ia sangat agresif dalam mempertahankan wilayahnya, bahkan berani menyerang elang sekalipun. Suaranya yang keras dan bervariasi mudah dikenali.",
    "Passer montanus": "Burung Gereja Erasia adalah spesies burung pipit yang sangat umum dan mudah dikenali di seluruh dunia. Dengan kepala cokelat dan bercak hitam di pipi, ia hidup berdampingan dengan manusia di lingkungan kampung dan pertanian.",
}

items = []
for row in rows:
    nama = row["nama_ilmiah"]
    foto = cari_foto(nama)
    desc = DESC.get(nama, (
        f"{row['nama_lokal']} ({nama}) adalah spesies burung dari famili {row['famili']} "
        f"yang berhasil diidentifikasi di sekitar Kampung Saukabu dan Pulau Fam, Raja Ampat. "
        f"Menurut daftar merah IUCN, spesies ini berstatus {row['iucn']}."
    ))
    items.append({
        "nama_ilmiah": nama,
        "nama_lokal": row["nama_lokal"],
        "famili": row["famili"],
        "iucn": row["iucn"],
        "cites": row["cites"],
        "p106": row["p106"],
        "deskripsi": desc,
        "image": foto
    })
    status = "✅" if foto else "❌ (no photo)"
    print(f"  [{row['no']:02d}] {nama} — {status}")

with open("data/fauna.json", "w", encoding="utf-8") as f:
    json.dump({"items": items}, f, ensure_ascii=False, indent=2)
print(f"\n✅ fauna.json: {len(items)} spesies\n")


# =================== FLORA ===================
print("=== MEMBANGUN FLORA.JSON ===")

POHON_DESC = {
    "Areca catechu L": {
        "nama_lokal": "Pinang", "famili": "Arecaceae",
        "desc": "Pinang adalah pohon palem langsing yang sangat erat kaitannya dengan budaya Papua. Bijinya dikunyah bersama sirih dan kapur sebagai tradisi sosial yang kuat di seluruh Papua, termasuk Saukabu. Pohon ini juga memiliki nilai ekonomi tinggi sebagai bahan baku industri."
    },
    "Artocarpus altilis": {
        "nama_lokal": "Sukun", "famili": "Moraceae",
        "desc": "Sukun adalah pohon penghasil buah bertepung yang menjadi sumber pangan penting di kepulauan Pasifik dan Papua. Buahnya yang besar dapat dibakar, direbus, atau digoreng. Di Saukabu, pohon sukun tumbuh di halaman warga dan menjadi sumber karbohidrat alternatif."
    },
    "Calophyllum inophyllum": {
        "nama_lokal": "Nyamplung / Bintangur", "famili": "Calophyllaceae",
        "desc": "Nyamplung adalah pohon pantai berbatang besar yang sangat khas di pesisir tropik. Kayunya kuat dan tahan air, banyak digunakan sebagai bahan bangunan dan pembuatan perahu. Bijinya mengandung minyak yang berpotensi sebagai bahan bakar nabati."
    },
    "Casuarina equisetifolia": {
        "nama_lokal": "Cemara Laut / Ru", "famili": "Casuarinaceae",
        "desc": "Cemara Laut adalah pohon pantai ikonik dengan daun seperti jarum yang melambai ditiup angin laut. Pohon ini sangat penting dalam menstabilkan garis pantai dari erosi gelombang. Di Saukabu, barisan cemara laut menjadi pelindung alami pemukiman dari hantaman ombak."
    },
    "Ceriops tagal": {
        "nama_lokal": "Tengar", "famili": "Rhizophoraceae",
        "desc": "Tengar adalah spesies mangrove dari famili bakau yang tumbuh di zona intertidal. Kulit kayunya mengandung tanin yang digunakan secara tradisional untuk menyamak kulit dan mengawetkan jaring ikan. Hutan mangrove di sekitar Saukabu merupakan ekosistem penting sebagai tempat pemijahan ikan."
    },
    "Cocos nucifera": {
        "nama_lokal": "Kelapa", "famili": "Arecaceae",
        "desc": "Kelapa adalah pohon serbaguna yang disebut pohon kehidupan di kepulauan tropis. Di Saukabu, kelapa hadir di mana-mana — buahnya diminum, dagingnya dimakan, minyaknya dimasak, daunnya dianyam, dan batangnya digunakan sebagai bahan bangunan. Kelapa muda gratis sering disajikan kepada tamu homestay."
    },
    "Ficus ampelas": {
        "nama_lokal": "Ampelas / Ara Ampelas", "famili": "Moraceae",
        "desc": "Ficus ampelas adalah pohon ara berukuran sedang dengan daun yang permukaannya kasar seperti ampelas, sehingga digunakan secara tradisional sebagai ampelas alami untuk menghaluskan kayu dan bahan lainnya. Spesies ini tumbuh di hutan sekunder dan tepi sungai."
    },
    "Hibiscus tiliaceus": {
        "nama_lokal": "Waru Laut", "famili": "Malvaceae",
        "desc": "Waru Laut adalah pohon pantai dengan bunga kuning cantik yang mekar di pagi hari. Kayunya dikenal sangat fleksibel dan kuat, sangat populer untuk bahan dayung, gagang alat, dan rangka perahu tradisional. Di Saukabu, waru tumbuh di sepanjang tepi pantai dan sungai."
    },
    "Mangifera indica L": {
        "nama_lokal": "Mangga", "famili": "Anacardiaceae",
        "desc": "Mangga adalah pohon buah tropis yang sangat populer di seluruh dunia. Di Saukabu, pohon mangga tumbuh di halaman warga dan menghasilkan buah yang dinikmati sebagai buah segar maupun olahan. Daun mudanya juga digunakan dalam berbagai tradisi adat setempat."
    },
    "Metroxylon sagu": {
        "nama_lokal": "Sagu", "famili": "Arecaceae",
        "desc": "Sagu adalah palem penghasil tepung yang menjadi makanan pokok utama masyarakat Papua, termasuk Saukabu. Setiap batang sagu dewasa dapat menghasilkan ratusan kilogram tepung sagu. Pohon ini hanya berbuah sekali seumur hidupnya (monokarpal) setelah bertahun-tahun tumbuh di rawa-rawa."
    },
    "Morinda citrifolia L": {
        "nama_lokal": "Mengkudu / Pace", "famili": "Rubiaceae",
        "desc": "Mengkudu adalah tanaman obat tradisional yang telah digunakan selama ribuan tahun oleh masyarakat Pasifik. Buahnya yang berbau khas mengandung senyawa bioaktif yang diyakini berkhasiat sebagai imunostimulan dan antioksidan. Di Saukabu, mengkudu tumbuh liar di sekitar pemukiman."
    },
    "Pandanus tectorius": {
        "nama_lokal": "Pandan Tikar / Pandan Laut", "famili": "Pandanaceae",
        "desc": "Pandan Tikar adalah tanaman pantai ikonik dengan batang bercabang-cabang dan akar tunjang yang kuat. Daunnya yang panjang dan kuat menjadi bahan utama anyaman tikar, topi, dan keranjang tradisional. Buahnya yang bulat menyerupai nanas juga dapat dimakan."
    },
    "Rhiozpora mucronata": {
        "nama_lokal": "Bakau Besar", "famili": "Rhizophoraceae",
        "desc": "Bakau Besar adalah spesies mangrove dominan di kawasan pesisir tropis. Akar tunjangnya yang kompleks berfungsi sebagai habitat nursery bagi ratusan spesies ikan dan invertebrata laut. Hutan mangrove di Saukabu sangat penting untuk menjaga kualitas air dan melindungi terumbu karang dari sedimentasi."
    },
    "Rhizopora stylosa": {
        "nama_lokal": "Bakau Kecil", "famili": "Rhizophoraceae",
        "desc": "Bakau Kecil adalah mangrove berukuran lebih kecil yang tumbuh di zona terdepan hutan mangrove, langsung berhadapan dengan ombak laut. Sistem akarnya yang padat menjadi benteng pertama dalam melindungi garis pantai Saukabu dari abrasi."
    },
    "Sonneratia alba": {
        "nama_lokal": "Bogem / Pedada", "famili": "Lythraceae",
        "desc": "Bogem atau Pedada adalah pohon mangrove berbunga putih yang mekar di malam hari dan diserbuki oleh kelelawar. Buahnya yang ranum dapat dimakan dan memiliki rasa asam-segar. Spesies ini tumbuh di zona mangrove terdepan yang tergenang air laut setiap hari."
    },
    "bambusa vulgaris": {
        "nama_lokal": "Bambu Kuning / Bambu Ampel", "famili": "Poaceae",
        "desc": "Bambu Ampel adalah salah satu spesies bambu yang paling banyak dimanfaatkan di dunia. Batangnya yang kuat namun ringan digunakan sebagai bahan bangunan, perabot, dan kerajinan tangan. Di Saukabu, bambu sering digunakan untuk membangun bagian-bagian rumah dan dermaga sederhana."
    },
    "bruguiera gymnorhiza": {
        "nama_lokal": "Tanjang / Bakau Tancang", "famili": "Rhizophoraceae",
        "desc": "Tancang adalah spesies mangrove dengan lutut akar (pneumatofora) yang menonjol dari lumpur dan berfungsi sebagai alat pernapasan. Kayunya sangat keras dan tahan terhadap air asin, sehingga sangat disukai sebagai bahan bakar dan material bangunan tradisional."
    },
    "ficus benjamina": {
        "nama_lokal": "Beringin / Ara Benjamin", "famili": "Moraceae",
        "desc": "Beringin Benjamin adalah pohon peneduh yang sangat populer. Di lingkungan alami, ia dapat tumbuh menjadi pohon raksasa dengan akar udara yang menjuntai. Buah-buah kecilnya yang matang menjadi sumber makanan penting bagi puluhan spesies burung, termasuk banyak spesies yang tercatat di Saukabu."
    },
    "heliotopium arboreum": {
        "nama_lokal": "Kangkung Laut / Heliotrope Pantai", "famili": "Boraginaceae",
        "desc": "Heliotrope Pantai adalah semak pantai dengan bunga putih kecil yang harum. Tanaman ini sangat toleran terhadap kondisi garam dan angin laut, sehingga sering dijumpai di barisan paling depan vegetasi pantai berpasir di Saukabu. Akarnya membantu mengikat pasir dari erosi angin."
    },
    "pongamia pinnata": {
        "nama_lokal": "Mempari / Kelor Laut", "famili": "Fabaceae",
        "desc": "Mempari adalah pohon pantai berukuran sedang dengan bunga ungu-merah muda yang harum dan polong berbiji tunggal. Bijinya mengandung minyak yang telah lama digunakan sebagai obat tradisional dan bahan bakar lampu. Pohon ini juga efektif mengikat nitrogen dan menyuburkan tanah."
    },
    "tectona grandis": {
        "nama_lokal": "Jati", "famili": "Lamiaceae",
        "desc": "Jati adalah kayu komersial paling berharga di dunia, dikenal dengan kekerasan, keindahan serat, dan ketahanan alaminya terhadap rayap dan cuaca. Keberadaan pohon jati di Saukabu kemungkinan merupakan hasil penanaman oleh warga untuk dimanfaatkan kayunya di masa depan."
    },
    "terminalia catappa": {
        "nama_lokal": "Ketapang / Kacang Laut", "famili": "Combretaceae",
        "desc": "Ketapang adalah pohon pantai ikonik dengan tajuk bertingkat-tingkat yang khas dan daun besar yang berubah merah sebelum gugur. Biji di dalam buahnya dapat dimakan dan rasanya mirip almond. Pohon ketapang menjadi peneduh favorit di tepi pantai Saukabu dan habitat penting bagi banyak spesies burung."
    },
}

flora_items = []
pohon_base = "images/Flora/Pohon"
for folder in sorted(os.listdir(pohon_base)):
    folder_path = os.path.join(pohon_base, folder)
    if not os.path.isdir(folder_path): continue
    
    # ambil foto pertama dari folder
    foto_path = ""
    for f in sorted(os.listdir(folder_path)):
        if f.lower().endswith(('.jpg','.jpeg','.png','.webp')):
            foto_path = f"{pohon_base}/{folder}/{f}".replace("\\", "/")
            break
    
    info = POHON_DESC.get(folder, {})
    flora_items.append({
        "nama_ilmiah": folder,
        "nama_lokal": info.get("nama_lokal", folder),
        "famili": info.get("famili", ""),
        "deskripsi": info.get("desc", f"{folder} adalah spesies tumbuhan yang ditemukan di Kampung Saukabu, Pulau Fam, Raja Ampat."),
        "image": foto_path
    })
    print(f"  ✅ {folder} — foto: {'ada' if foto_path else 'tidak ada'}")

with open("data/flora.json", "w", encoding="utf-8") as f:
    json.dump({"items": flora_items}, f, ensure_ascii=False, indent=2)
print(f"\n✅ flora.json: {len(flora_items)} spesies tumbuhan")
print("\nSelesai! Jalankan git add -A && git commit -m 'update fauna flora' && git push origin main")
