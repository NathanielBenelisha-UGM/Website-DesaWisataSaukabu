"""
Skrip untuk mengekstrak data burung dari file Excel "Data burung saukabu.xlsx"
Akan membuat:
  1. images/InventarisBurung/ - folder berisi foto-foto burung
  2. data/fauna.json - file data burung untuk website

Cara pakai:
  1. Pastikan Python sudah terinstall
  2. Install openpyxl: pip install openpyxl
  3. Jalankan skrip ini dari folder E:\KKN-WEBSITE_SAUKABU
"""

import json
import os
import zipfile
import shutil
from openpyxl import load_workbook

# === KONFIGURASI ===
EXCEL_FILE = "Data burung saukabu.xlsx"
OUTPUT_IMAGE_DIR = "images/InventarisBurung"
OUTPUT_JSON = "data/fauna.json"

# Deskripsi valid per spesies (akan digunakan jika ada di dict ini)
# Tambahkan deskripsi manual di sini jika diperlukan
DESKRIPSI_MANUAL = {}

# Pastikan folder output ada
os.makedirs(OUTPUT_IMAGE_DIR, exist_ok=True)

print(f"Membaca file: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# Baca data dari kolom A-H, mulai baris 2 (baris 1 = header)
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    no, nama_ilmiah, nama_lokal, famili, iucn, cites, p106, gambar = (row + (None,)*8)[:8]
    if not nama_ilmiah:
        continue
    rows.append({
        "no": no,
        "nama_ilmiah": str(nama_ilmiah).strip() if nama_ilmiah else "",
        "nama_lokal": str(nama_lokal).strip() if nama_lokal else "",
        "famili": str(famili).strip() if famili else "",
        "iucn": str(iucn).strip() if iucn else "LC",
        "cites": str(cites).strip() if cites else "-",
        "p106": str(p106).strip() if p106 else "-",
    })

print(f"  Ditemukan {len(rows)} spesies burung")

# Ekstrak gambar yang tertanam di dalam file Excel
print("\nMengekstrak foto dari Excel...")
image_map = {}  # Untuk memetakan gambar ke baris data

# Excel sebenarnya adalah file ZIP, kita buka langsung
with zipfile.ZipFile(EXCEL_FILE, 'r') as z:
    media_files = [f for f in z.namelist() if f.startswith('xl/media/')]
    print(f"  Ditemukan {len(media_files)} file gambar di dalam Excel")
    
    for i, media_path in enumerate(sorted(media_files)):
        ext = media_path.split('.')[-1]
        # Ambil nama burung dari data row (berdasarkan urutan)
        if i < len(rows):
            nama_ilmiah = rows[i]["nama_ilmiah"].replace(" ", "_").replace("/", "-")
            nama_file = f"{nama_ilmiah}.{ext}"
        else:
            nama_file = f"burung_{i+1}.{ext}"
        
        output_path = os.path.join(OUTPUT_IMAGE_DIR, nama_file)
        with z.open(media_path) as src, open(output_path, 'wb') as dst:
            dst.write(src.read())
        
        image_map[i] = nama_file
        print(f"  [{i+1}] Tersimpan: {nama_file}")

# Bangun deskripsi otomatis berdasarkan data yang ada
def buat_deskripsi(row):
    nama = row["nama_ilmiah"]
    nama_lokal = row["nama_lokal"]
    famili = row["famili"]
    iucn = row["iucn"]
    
    # Deskripsi per status IUCN
    status_desc = {
        "LC": "berstatus Least Concern (Risiko Rendah)",
        "NT": "berstatus Near Threatened (Hampir Terancam)",
        "VU": "berstatus Vulnerable (Rentan)",
        "EN": "berstatus Endangered (Terancam)",
        "CR": "berstatus Critically Endangered (Kritis)",
        "DD": "berstatus Data Deficient (Data Kurang)",
    }
    status_text = status_desc.get(iucn, f"berstatus {iucn}")
    
    if nama in DESKRIPSI_MANUAL:
        return DESKRIPSI_MANUAL[nama]
    
    return (f"{nama_lokal} ({nama}) adalah spesies burung dari famili {famili} "
            f"yang berhasil didokumentasikan di sekitar Kampung Saukabu. "
            f"Berdasarkan daftar merah IUCN, spesies ini {status_text}.")

# Buat fauna.json
items = []
for i, row in enumerate(rows):
    nama_ilmiah = row["nama_ilmiah"].replace(" ", "_").replace("/", "-")
    
    # Cari file gambar yang sesuai
    image_path = ""
    for ext in ["jpeg", "jpg", "png", "webp"]:
        candidate = f"images/InventarisBurung/{nama_ilmiah}.{ext}"
        if os.path.exists(candidate):
            image_path = candidate
            break
    # Fallback ke peta gambar berdasarkan urutan
    if not image_path and i in image_map:
        image_path = f"images/InventarisBurung/{image_map[i]}"
    
    item = {
        "nama_ilmiah": row["nama_ilmiah"],
        "nama_lokal": row["nama_lokal"],
        "famili": row["famili"],
        "iucn": row["iucn"],
        "cites": row["cites"],
        "p106": row["p106"],
        "deskripsi": buat_deskripsi(row),
        "image": image_path
    }
    items.append(item)

fauna_data = {"items": items}

with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(fauna_data, f, ensure_ascii=False, indent=2)

print(f"\n✅ Selesai!")
print(f"   - {len(items)} burung tersimpan di {OUTPUT_JSON}")
print(f"   - Foto tersimpan di folder {OUTPUT_IMAGE_DIR}/")
print(f"\nSilakan buka website untuk melihat hasilnya.")
