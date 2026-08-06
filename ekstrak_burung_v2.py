"""
Skrip v2 - Ekstrak 39 burung dari Excel + deskripsi lengkap + fallback Wikimedia
Jalankan: python ekstrak_burung_v2.py
"""
import json, os, zipfile
from openpyxl import load_workbook

EXCEL_FILE = "Data burung saukabu.xlsx"
OUTPUT_IMAGE_DIR = "images/fauna"
OUTPUT_JSON = "data/fauna.json"
os.makedirs(OUTPUT_IMAGE_DIR, exist_ok=True)

# Deskripsi lengkap per nama ilmiah
DESKRIPSI = {
    "Ptilinopus rivoli": "Walik Dada Putih adalah merpati buah paling elegan di Raja Ampat. Jantan dewasa memiliki bulu hijau zamrud berkilau dengan dada putih bersih yang kontras. Betina umumnya hijau seluruhnya. Spesies ini menghuni hutan primer dan tepi hutan, memakan buah dan beri dari pohon tinggi. Di Saukabu, ia sering terdengar sebelum terlihat — suara panggilannya yang berulang menjadi penanda kehadiran hutan sehat.",
    "Eclectus roratus": "Burung Bayan adalah salah satu spesies beo paling mencolok di dunia, dengan perbedaan warna jantan-betina yang ekstrem. Jantan berwarna hijau cerah dengan paruh merah-oranye, betina berwarna merah tua dengan perut biru dan paruh hitam. Di Saukabu, keduanya sering terlihat berpasangan di tajuk pohon tinggi. Suara panggilannya yang keras menjadi ikon bunyi hutan Papua.",
    "Geoffroyus geoffroyi": "Nuri Pipi-Merah adalah nuri berukuran sedang yang sangat aktif. Jantan memiliki kepala merah muda keunguan yang khas, betina berkepala cokelat. Keduanya bertubuh hijau dominan. Spesies ini berasosiasi kuat dengan hutan primer yang masih baik. Kelompok kecil Nuri Pipi-Merah kerap terlihat terbang cepat di antara kanopi pohon buah di sekitar Kampung Saukabu.",
    "Rhipidura leucophrys": "Kipasan Belang adalah burung kecil yang aktif dan tidak takut manusia. Dengan bulu hitam-putih kontras dan ekor yang selalu dikibas-kibaskan, ia menjadi teman setia saat berjalan di sekitar kampung. Spesies ini memangsa serangga kecil di udara dengan manuver terbang yang akrobatik. Kehadirannya menjadi indikator ekosistem yang masih terjaga.",
    "Tanysiptera galatea": "Cekakak Sega adalah kingfisher bermahkota biru cerah dengan ekor panjang berwarna putih yang sangat khas. Spesies yang cantik ini menghuni hutan dataran rendah dan tepi sungai. Di Raja Ampat, ia termasuk spesies yang ikonik dan menjadi daya tarik bagi pengamat burung dari seluruh dunia.",
    "Halcyon chloris": "Cekakak Sungai adalah kingfisher yang paling umum dijumpai di pesisir dan tepi sungai. Bulu biru kehijauan di punggung dan mahkota biru tua menjadikannya mudah dikenali. Spesies ini merupakan predator aktif ikan-ikan kecil, katak, dan serangga. Kerap terlihat bertengger diam di ranting yang menggantung di atas air.",
    "Todiramphus sanctus": "Cekakak Suci adalah kingfisher berukuran sedang dengan punggung biru-hijau mengkilap dan dada putih bersih. Spesies migran parsial ini sering terlihat di tepi pantai, mangrove, dan kebun kelapa di sekitar Saukabu. Ia merupakan salah satu burung pesisir yang paling mudah dijumpai di kepulauan Raja Ampat.",
    "Alcedo pusilla": "Raja Udang Kerdil adalah kingfisher terkecil di Papua. Dengan ukuran yang hanya sebesar jempol, bulu biru zamrud berkilaunya tetap mencolok. Spesies ini sangat pemalu dan menghuni tepian sungai kecil serta saluran air di dalam hutan. Penampakan langsung spesies ini adalah pencapaian tersendiri bagi pengamat burung.",
    "Nectarinia jugularis": "Burung Madu Sriganti adalah burung madu yang paling umum di Asia Tenggara dan Papua. Jantan memiliki bulu atas hijau mengkilap dan tenggorokan merah-ungu yang indah. Betina berwarna lebih polos kekuningan. Ia memakan nektar bunga dan serangga kecil. Di Saukabu, ia aktif di pagi hari mengunjungi pohon-pohon berbunga di sekitar kampung.",
    "Cinnyris sericeus": "Burung Madu Polos atau Sriganti Laut adalah burung madu dengan tubuh ramping dan paruh melengkung panjang. Jantan dewasa berwarna hitam mengkilap dengan kilau ungu-biru pada kepala dan punggung. Spesies ini menyukai habitat tepi pantai, mangrove, dan hutan dataran rendah di kepulauan Papua.",
    "Lorius lory": "Nuri Bayan Raja adalah salah satu burung paling mencolok di Papua. Dengan dominasi merah terang, sayap hijau, dan detail biru-hitam yang kaya, spesies ini benar-benar memukau. Ia menghuni hutan primer dataran rendah hingga pegunungan bawah. Di Raja Ampat, keberadaannya menjadi daya tarik utama bagi fotografer alam liar.",
    "Charmosyna papou": "Perkici Papuan atau Lori Papuan adalah lori kecil berwarna merah, hijau, dan kuning yang sangat cantik. Ekornya yang panjang dan lancip menjadikannya mudah dikenali dalam penerbangan. Spesies ini memakan nektar dan serbuk sari bunga. Ia aktif di pagi hari dan sering terlihat berpasangan atau dalam kelompok kecil.",
    "Cacatua goffini": "Kakatua Tanimbar (Kakatua Gofin) adalah kakatua putih berukuran kecil yang endemik kepulauan Tanimbar dan kini juga tercatat di Raja Ampat sebagai hasil introduksi. Spesies ini berstatus Rentan (VU) dalam daftar IUCN karena ancaman perdagangan ilegal. Kakatua ini dikenal cerdas dan memiliki kemampuan pemecahan masalah yang luar biasa.",
    "Probosciger aterrimus": "Kakatua Raja adalah kakatua hitam terbesar dan paling ikonik di Papua. Dengan jambul panjang yang mengesankan dan pipi merah tanpa bulu yang unik, spesies ini sangat mudah dikenali. Paruhnya yang sangat kuat mampu memecah biji keras. Kakatua Raja berstatus dilindungi penuh dan menjadi simbol kebanggaan satwa liar Papua.",
    "Ducula spilorrhoa": "Pergam Pala adalah merpati besar pemakan buah yang sangat penting bagi regenerasi hutan. Dengan bulu abu-abu pucat, sayap hijau gelap, dan ekor bergaris hitam, ia mudah dikenali di tajuk pohon. Spesies ini memiliki peran krusial sebagai pemencar biji buah-buahan hutan di seluruh kepulauan Papua.",
    "Ptilinopus superbus": "Walik Superba atau Merpati Buah Megah adalah merpati buah yang sangat indah. Jantan memiliki kepala ungu-biru, punggung hijau zamrud, dan perut oranye-merah yang kontras. Betina lebih polos kehijauan. Spesies ini menghuni hutan primer dan tepi hutan di seluruh Papua, memakan berbagai jenis buah-buahan.",
    "Ptilinopus magnificus": "Walik Cikalong adalah merpati buah berukuran besar dengan tampilan yang sangat mencolok. Kepala kuning cerah, punggung hijau, dan dada oranye-merah menjadikannya salah satu merpati buah terindah di dunia. Ia menghuni hutan tropis primer di dataran rendah Papua dan kepulauan di sekitarnya.",
    "Macropygia amboinensis": "Uncal Ambon adalah merpati berukuran sedang dengan tubuh ramping dan ekor panjang bertingkat. Berwarna cokelat kemerahan di atas dengan dada merah muda pucat. Spesies ini menyukai tepi hutan, semak belukar, dan kebun. Di Saukabu, ia sering terdengar berkukuk lembut di pagi hari dari balik pepohonan.",
    "Megapodius freycinet": "Maleo Gosong adalah megapoda berukuran ayam yang tidak mengerami telurnya sendiri, melainkan mengubur telur di dalam gundukan pasir atau tanah hangat. Panas bumi dan panas matahari menggantikan peran induk dalam proses penetasan. Spesies ini adalah insinyur alam yang luar biasa dan sangat penting bagi keanekaragaman hayati Papua.",
    "Egretta sacra": "Kuntul Karang adalah kuntul yang hidup di ekosistem pesisir batu karang dan pantai berbatu. Terdapat dua morfa: putih seluruhnya, atau abu-abu gelap seluruhnya. Ia adalah predator aktif ikan dan crustasea di zona pasang surut. Sering terlihat berdiri diam di atas batu karang yang diperciki ombak sambil mengincar mangsa.",
    "Egretta garzetta": "Kuntul Kecil adalah kuntul serba putih berukuran sedang dengan paruh dan kaki hitam serta jari kaki berwarna kuning cerah. Spesies kosmopolitan ini umum dijumpai di kawasan pesisir, rawa, dan sawah. Di Saukabu, ia terlihat mencari makan di perairan dangkal pantai saat air surut.",
    "Ardea purpurea": "Cangak Merah adalah cangak berukuran besar dengan leher dan badan berwarna merah kecokelatan bergaris hitam. Ia lebih menyukai rawa-rawa bervegetasi lebat dan tepi sungai. Spesies pemalu ini jarang terlihat terbang di siang hari dan lebih sering bersembunyi di balik vegetasi rawa yang rapat.",
    "Butorides striata": "Kokokan Laut adalah cangak kecil yang sangat umum di pesisir dan mangrove. Dengan bulu abu-abu-hijau di atas dan kekuningan di bawah, ia menyesuaikan diri dengan sangat baik di ekosistem pesisir. Spesies ini adalah pemburu oportunistik yang memakan ikan, udang, dan serangga di tepi air.",
    "Fregata minor": "Cikalang Besar adalah burung laut yang sangat mahir terbang, dengan rentang sayap lebih dari dua meter. Spesies ini tidak bisa mendarat di air karena bulunya tidak tahan air, sehingga ia menangkap ikan langsung dari permukaan laut atau merampas dari burung lain (kleptoparasitisme). Kantung gular merah pada jantan menjadi daya tarik musim kawin yang menakjubkan.",
    "Haliastur indus": "Elang Bondol adalah elang berukuran sedang dengan tubuh putih bersih dan sayap cokelat kemerahan. Spesies ini adalah elang pesisir yang sangat mudah dijumpai di seluruh nusantara. Di Saukabu, ia sering terlihat melayang di atas pantai dan perairan, mencari ikan, crustasea, dan bangkai hewan laut.",
    "Pandion haliaetus": "Elang Tiram adalah elang spesialis pemangsa ikan yang tersebar di seluruh dunia. Ia memiliki adaptasi kaki yang unik — dua jari menghadap depan, dua ke belakang — untuk mencengkeram ikan yang licin. Dikenal menyelam kaki-pertama ke dalam air untuk menangkap ikan. Di Raja Ampat, spesies migran ini terlihat terutama saat musim migrasi.",
    "Accipiter novaehollandiae": "Elang Alap Abu-abu adalah accipiter berukuran sedang dengan punggung abu-abu dan dada putih bergaris oranye halus. Spesies ini merupakan predator burung-burung kecil dan reptil di hutan dataran rendah. Penerbangannya yang cepat dan lincah di antara pepohonan menjadikan ia pemburu yang sangat efektif.",
    "Falco peregrinus": "Alap-alap Kawah atau Elang Peregrine adalah burung pemangsa tercepat di dunia, mampu menukik dengan kecepatan lebih dari 300 km/jam saat memburu mangsa. Spesies migran ini terlihat di Raja Ampat terutama saat musim gugur dan musim semi. Kehadiran Elang Peregrine menunjukkan kekayaan ekosistem burung di kawasan ini.",
    "Spilornis cheela": "Elang Ular Bido adalah elang yang terspesialisasi memangsa ular dan reptil lain. Jambul pendek yang khas dan bagian bawah berbintik putih menjadikannya mudah dikenali. Di Saukabu, ia kerap terlihat melayang tinggi di udara di atas tepian hutan, sambil mengeluarkan teriakan nyaring yang khas.",
    "Cacomantis variolosus": "Wiwik Kelabu adalah sejenis kukuk berukuran kecil hingga sedang. Seperti kebanyakan anggota familinya, wiwik bersifat parasit sarang — ia menitipkan telurnya ke sarang burung lain untuk diierami dan dibesarkan. Suaranya yang berulang dan melengking sering terdengar di sore dan malam hari di tepi hutan Saukabu.",
    "Eudynamys scolopaceus": "Tuwur Asia atau Koels adalah burung dari famili kukuk yang dikenal dengan suaranya yang keras dan berulang, makin lama makin cepat. Jantan berwarna hitam mengkilap, betina berwarna cokelat bergaris. Spesies ini juga bersifat parasit sarang, terutama pada sarang gagak. Suaranya menjadi penanda musim panas di berbagai bagian Asia dan Papua.",
    "Centropus violaceus": "Bubut Ungu adalah spesies bubut yang hidup di hutan dataran rendah dan hutan pantai. Berbeda dengan kebanyakan kukuk, bubut membangun sarangnya sendiri dan merawat anaknya. Berwarna hitam-keunguan mengkilap dengan bagian bawah cokelat karat. Spesies yang cenderung pemalu dan lebih sering terdengar suaranya daripada terlihat tubuhnya.",
    "Collocalia esculenta": "Walet Linchi adalah walet kecil yang sangat umum. Dengan tubuh hitam legam mengkilap di atas dan bagian bawah putih, ia mudah dikenali. Walet ini membangun sarang dari rumput yang dicampur air liurnya. Sarang dari genus Collocalia inilah yang menjadi bahan sup sarang burung yang terkenal. Koloni walet terlihat terbang berputar-putar di langit Saukabu setiap sore.",
    "Todiramphus macleayii": "Cekakak Hutan adalah kingfisher berukuran sedang dengan punggung biru cerah dan perut putih bersih. Ciri khasnya adalah kerah putih yang lebar di leher. Spesies ini menyukai hutan terbuka, tepi hutan, dan kebun. Di Raja Ampat, ia adalah salah satu kingfisher yang paling mudah dijumpai di habitat daratan.",
    "Pitta erythrogaster": "Paok Dada Merah adalah burung paok yang sangat cantik dengan warna-warna mencolok: punggung hijau, kepala hitam, dada dan perut merah terang. Ia adalah burung hutan lantai yang pemalu dan sulit dilihat meski suaranya yang nyaring sering terdengar. Keberadaannya menjadi indikator kesehatan hutan primer dataran rendah.",
    "Corvus orru": "Gagak Torresian adalah gagak hitam legam berukuran besar yang sangat cerdas dan adaptif. Ia adalah omnivora oportunistik yang memakan hampir segalanya: bangkai, serangga, telur burung lain, buah, dan sampah. Di Saukabu, gagak terlihat berkelompok dan sering berinteraksi dengan aktivitas warga kampung.",
    "Dicrurus bracteatus": "Srigunting Perahu adalah burung hitam mengkilap berukuran sedang dengan ekor bercabang seperti garpu yang khas. Ia adalah pemburu serangga yang sangat agresif dan sering menyerang burung yang lebih besar — termasuk elang — untuk melindungi wilayahnya. Suaranya yang keras dan bervariasi menjadikannya penghuni hutan yang mudah dikenali.",
    "Mino dumontii": "Jalak Emas Papua adalah burung jalak berkilau dengan bulu hitam, kuning emas mencolok di sayap dan ekor, serta kulit wajah kuning tanpa bulu. Spesies ini sangat vokal dan hidup berkelompok. Di Raja Ampat, Jalak Emas adalah salah satu burung yang paling mudah dan paling menggembirakan untuk dijumpai di hutan.",
    "Aplonis metallica": "Perling Sutera adalah burung jalak berukuran sedang dengan bulu hitam mengkilap berwarna hijau-ungu metalik saat terkena cahaya. Spesies ini hidup berkoloni besar dan membangun sarang menggantung di pohon-pohon tinggi, sering kali dalam jumlah ratusan sarang dalam satu pohon. Di Saukabu, koloninya menjadi pemandangan yang ramai dan ribut setiap sore.",
    "Lonchura spectabilis": "Bondol Papuan adalah burung pipit kecil yang cantik dengan kepala hitam dan tubuh cokelat-merah tua. Ia hidup berkelompok dan memakan biji-bijian rumput serta padi. Spesies ini sangat umum dijumpai di ladang, semak belukar, dan tepi kampung di seluruh Papua. Di Saukabu, kelompok bondol kerap terlihat mencari makan di tepi jalan dan kebun warga.",
}

# Wikimedia fallback untuk yang tidak punya foto
WIKIMEDIA_FALLBACK = {
    "Tanysiptera galatea": "https://upload.wikimedia.org/wikipedia/commons/thumb/a/a3/Tanysiptera_galatea_-Vogelkop-8.jpg/800px-Tanysiptera_galatea_-Vogelkop-8.jpg",
    "Halcyon chloris": "https://upload.wikimedia.org/wikipedia/commons/thumb/7/7f/Collared_kingfisher_perching.jpg/800px-Collared_kingfisher_perching.jpg",
    "Todiramphus sanctus": "https://upload.wikimedia.org/wikipedia/commons/thumb/b/b1/Sacred_kingfisher_Feb09.jpg/800px-Sacred_kingfisher_Feb09.jpg",
    "Alcedo pusilla": "https://upload.wikimedia.org/wikipedia/commons/thumb/f/f3/Little_Kingfisher.jpg/800px-Little_Kingfisher.jpg",
    "Nectarinia jugularis": "https://upload.wikimedia.org/wikipedia/commons/thumb/3/37/Nectarinia_jugularis_-_Thailand.jpg/800px-Nectarinia_jugularis_-_Thailand.jpg",
}

print(f"Membaca file: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cols = (row + (None,)*8)[:8]
    no, nama_ilmiah, nama_lokal, famili, iucn, cites, p106, gambar = cols
    if not nama_ilmiah:
        continue
    rows.append({
        "no": no,
        "nama_ilmiah": str(nama_ilmiah).strip(),
        "nama_lokal": str(nama_lokal).strip() if nama_lokal else "",
        "famili": str(famili).strip() if famili else "",
        "iucn": str(iucn).strip() if iucn else "LC",
        "cites": str(cites).strip() if cites else "-",
        "p106": str(p106).strip() if p106 else "-",
    })
print(f"  {len(rows)} spesies ditemukan")

# Ekstrak foto dari Excel
print("\nMengekstrak foto dari Excel...")
image_map = {}
with zipfile.ZipFile(EXCEL_FILE, 'r') as z:
    media_files = sorted([f for f in z.namelist() if f.startswith('xl/media/')])
    print(f"  {len(media_files)} foto ditemukan di Excel")
    for i, media_path in enumerate(media_files):
        ext = media_path.split('.')[-1].lower()
        if i < len(rows):
            nama = rows[i]["nama_ilmiah"].replace(" ", "_").replace("/", "-")
            nama_file = f"{nama}.{ext}"
        else:
            nama_file = f"burung_{i+1}.{ext}"
        out = os.path.join(OUTPUT_IMAGE_DIR, nama_file)
        with z.open(media_path) as src, open(out, 'wb') as dst:
            dst.write(src.read())
        image_map[i] = nama_file
        print(f"  [{i+1}] {nama_file}")

# Build fauna.json
items = []
for i, row in enumerate(rows):
    nama = row["nama_ilmiah"]
    nama_key = nama.replace(" ", "_").replace("/", "-")
    
    # Cari foto lokal
    image_path = ""
    for ext in ["jpeg", "jpg", "png", "webp"]:
        candidate = f"images/InventarisBurung/{nama_key}.{ext}"
        if os.path.exists(candidate):
            image_path = candidate
            break
    if not image_path and i in image_map:
        image_path = f"images/InventarisBurung/{image_map[i]}"
    
    # Fallback Wikimedia jika tidak ada foto
    if not image_path and nama in WIKIMEDIA_FALLBACK:
        image_path = WIKIMEDIA_FALLBACK[nama]
    
    desc = DESKRIPSI.get(nama, (
        f"{row['nama_lokal']} ({nama}) adalah spesies burung dari famili {row['famili']} "
        f"yang berhasil didokumentasikan di sekitar Kampung Saukabu, Pulau Fam, Raja Ampat. "
        f"Menurut daftar merah IUCN, spesies ini berstatus {row['iucn']}."
    ))
    
    items.append({
        "nama_ilmiah": nama,
        "nama_lokal": row["nama_lokal"],
        "famili": row["famili"],
        "iucn": row["iucn"],
        "cites": row["cites"],
        "p106": row["p106"],
        "deskripsi": desc,
        "image": image_path
    })

with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump({"items": items}, f, ensure_ascii=False, indent=2)

print(f"\n✅ Selesai! {len(items)} burung → {OUTPUT_JSON}")
print(f"   Foto lokal: {len(image_map)}, Fallback Wikimedia: tersedia untuk 5 spesies")
